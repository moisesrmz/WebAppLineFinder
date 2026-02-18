import pandas as pd
import backend.config as config 
import shutil
from datetime import datetime

def buscar_numero_parte(numeros_parte):
    mensaje = ""
    filas_resaltadas = []


    try:
        df1 = pd.read_excel(config.DATA_FILE, sheet_name="Hoja1", engine="openpyxl")
        df2 = pd.read_excel(config.DATA_FILE, sheet_name="Hoja2", engine="openpyxl")
        df3 = pd.read_excel(config.DATA_FILE, sheet_name="Hoja3", engine="openpyxl")
    except Exception as e:
        return f"Error al leer el archivo Excel: {str(e)}", []

    # Convertir a string y eliminar espacios extra
    df1 = df1.applymap(lambda x: str(x).strip() if pd.notna(x) else x)
    df2 = df2.applymap(lambda x: str(x).strip() if pd.notna(x) else x)
    df3 = df3.applymap(lambda x: str(x).strip() if pd.notna(x) else x)

    for numero_parte in numeros_parte:
        fila_df1 = df1[df1.iloc[:, 0].astype(str).str.strip() == numero_parte]

        if fila_df1.empty:
            mensaje += f"<div class='resultado'><h3>Número de parte '{numero_parte}' no encontrado en la base de datos.</h3></div>"
        else:
            resultados_ordenados = []

            for _, row in fila_df1.iterrows():
                molex_pn = row[0]
                version = row[1]   # Valor de la columna B
                valores_buscados = row[2:].dropna().tolist()
                
                # ✅ Contamos cuántas veces aparece cada valor en la búsqueda
                valores_buscados_contador = {val: valores_buscados.count(val) for val in set(valores_buscados)}
                
                resultados_coincidencias = set()

                for _, row_df2 in df2.iterrows():
                    valores_encontrados_contador = {}  # ✅ Contador de valores encontrados en esta fila

                    for value in row_df2.iloc[1:].dropna():
                        valores_encontrados_contador[value] = valores_encontrados_contador.get(value, 0) + 1

                    # ✅ Verificamos que cada valor buscado esté en la fila con la misma frecuencia
                    coincidencia_valida = all(
                        valores_encontrados_contador.get(val, 0) >= cantidad
                        for val, cantidad in valores_buscados_contador.items()
                    )

                    if coincidencia_valida:
                        resultados_coincidencias.add(row_df2.iloc[0])  # ✅ Guardamos el nombre del tester

                for tester in resultados_coincidencias:
                    fila_df3 = df3[df3.iloc[:, 0].astype(str).str.strip() == tester]

                    if not fila_df3.empty:
                        encontrado = any(numero_parte in str(cell) for cell in fila_df3.iloc[0, :])
                        if encontrado:
                            resultados_ordenados.append(f"{molex_pn} - {tester}: <span style='color:green; font-weight:bold;'>Confirmado</span>")
                        else:
                            resultados_ordenados.append(f"{molex_pn} - {tester}: <span style='color:red; font-weight:bold;'>Validar con Ingeniería de Pruebas en piso</span>")
                    else:
                        resultados_ordenados.append(f"{molex_pn} - {tester}: Tester no habilitado")

            if resultados_ordenados:
                resultados_ordenados = list(set(resultados_ordenados))
                resultados_ordenados.sort()
                # Usamos la variable 'version' como descripción en el título del resultado para mayor claridad
                mensaje += "<div class='resultado'><h3>Resultados para el número de parte '{0}' ({1}):</h3>".format(numero_parte, version)
                mensaje += f"<p><strong>Módulos:</strong> {', '.join(valores_buscados)}</p>"
                for resultado in resultados_ordenados:
                    mensaje += f"<br>{resultado}<br>"
                mensaje += "</div>"
            else:
                mensaje += (
                    "<div class='resultado'><h3>Resultados para el número de parte '{0}' ({1}):</h3>"
                    "<p><strong>No se encontró tester con módulos:</strong> {2}</p></div>"
                ).format(numero_parte, version, ', '.join(valores_buscados))



    return mensaje, filas_resaltadas
def buscar_modulo(modulo):
    mensaje = ""
    filas_resaltadas = []

    try:
        # ✅ Leer la hoja 2
        df = pd.read_excel(config.DATA_FILE, sheet_name="Hoja2", engine="openpyxl", dtype=str)
        df.fillna("", inplace=True)  # Evita errores por NaN
    except Exception as e:
        print(f" Error al leer Hoja2: {e}")
        return f"Error al leer el archivo Excel: {str(e)}", []

    # ✅ Convertir todo a minúsculas y quitar espacios
    df = df.applymap(lambda x: str(x).strip().lower() if pd.notna(x) else "")
    modulo = str(modulo).strip().lower()

    # ✅ Buscar coincidencias (en cualquier columna a partir de la 3)
    coincidencias = []
    for _, fila in df.iterrows():
        valores = [str(v) for v in fila[2:].tolist()]
        cantidad = sum(1 for v in valores if modulo in v)

        if cantidad > 0:
            coincidencias.append({
                "FA": fila.iloc[0],
                "EOL": fila.iloc[1],
                "Cantidad": f"{cantidad}x"
            })

    # ✅ Generar mensaje HTML
    if coincidencias:
        mensaje += f"<div class='resultado'><h3>Resultados para el módulo '{modulo}':</h3>"
        mensaje += "<table border='1'><tr><th>FA</th><th>EOL</th><th>Cantidad</th></tr>"
        for c in coincidencias:
            mensaje += f"<tr><td>{c['FA']}</td><td>{c['EOL']}</td><td>{c['Cantidad']}</td></tr>"
        mensaje += "</table></div>"
        filas_resaltadas = coincidencias
    else:
        mensaje += f"<div class='resultado'><h3>No se encontraron líneas con el módulo '{modulo}'.</h3></div>"

    print(f" Buscar módulo: '{modulo}', coincidencias encontradas: {len(coincidencias)}")
    return mensaje, filas_resaltadas

# Leer datos de una hoja del Excel
def leer_hoja(hoja):
    try:
        df = pd.read_excel(config.DATA_FILE, sheet_name=hoja, engine="openpyxl", dtype=str)  # ✅ Convertir todo a string
        df.fillna("", inplace=True)  # ✅ Reemplazar valores NaN con cadena vacía
        print(f"Hoja {hoja} cargada con {df.shape[0]} filas y {df.shape[1]} columnas")  # 👀 Depuración
        print(df.head().to_dict())  # 👀 Verificar datos en consola
        return df.values.tolist()  # Convertir DataFrame a lista de listas
    except Exception as e:
        print(f"Error al cargar hoja {hoja}: {e}")
        return {"error": str(e)}

# 🟢 Guardar cambios asegurando que se mantienen los encabezados
def guardar_cambios_excel(hoja, datos):
    try:
        df = pd.DataFrame(datos[1:], columns=datos[0])  # ✅ Mantiene encabezados

        with pd.ExcelWriter(config.DATA_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
            df.to_excel(writer, sheet_name=hoja, index=False)

        return "Guardado correctamente con encabezados"
    except Exception as e:
        return f"Error: {str(e)}"
# ============================================================
# 🚨 ALERTAS DE CALIDAD (PERSISTENTE EN /home)
# ============================================================

import os
import shutil
from datetime import datetime
from openpyxl import Workbook, load_workbook

ALERTAS_SHEET = "Alertas"

# 🔐 Ruta persistente real en Azure
BASE_HOME = "/home"

if not os.path.exists(BASE_HOME):
    # fallback para entorno local
    BASE_HOME = os.getcwd()

ALERTAS_DIR = os.path.join(BASE_HOME, "data")
BACKUP_DIR = os.path.join(BASE_HOME, "backups")

ALERTAS_FILE = os.path.join(ALERTAS_DIR, "AlertasCalidad.xlsx")

# Crear carpetas si no existen
os.makedirs(ALERTAS_DIR, exist_ok=True)
os.makedirs(BACKUP_DIR, exist_ok=True)

print("=== ALERTAS PATH ===")
print("BASE_HOME:", BASE_HOME)
print("ALERTAS_FILE:", ALERTAS_FILE)


def _inicializar_alertas():
    if not os.path.exists(ALERTAS_FILE):
        print("Creando archivo inicial de alertas...")
        wb = Workbook()
        ws = wb.active
        ws.title = ALERTAS_SHEET
        ws.append([
            "ID",
            "Numero_Parte_o_Area",
            "Numero_Alerta",
            "Descripcion",
            "Fecha_Generacion",
            "Vigente_Hasta"
        ])
        wb.save(ALERTAS_FILE)


def leer_alertas():
    _inicializar_alertas()

    print("Leyendo alertas desde:", ALERTAS_FILE)
    print("Existe archivo?:", os.path.exists(ALERTAS_FILE))

    wb = load_workbook(ALERTAS_FILE)
    ws = wb[ALERTAS_SHEET]

    alertas = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            alertas.append({
                "id": row[0],
                "numero_parte_area": str(row[1]) if row[1] else "",
                "numero_alerta": str(row[2]) if row[2] else "",
                "descripcion": str(row[3]) if row[3] else "",
                "fecha_generacion": str(row[4]) if row[4] else "",
                "vigente_hasta": str(row[5]) if row[5] else ""
            })

    return alertas


def guardar_alertas(data):

    _inicializar_alertas()

    # 🔒 Validación fuerte
    for alerta in data:
        campo = alerta["numero_parte_area"]

        if not (
            (campo.isdigit() and len(campo) == 10)
            or campo.strip().lower() in ["fa", "corte", "crimp"]
        ):
            raise ValueError(
                "El campo debe ser 10 dígitos o FA, Corte, Crimp"
            )

    print("Guardando alertas en:", ALERTAS_FILE)

    # 📦 BACKUP automático (guardado en carpeta separada)
    if os.path.exists(ALERTAS_FILE):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_name = f"AlertasCalidad_backup_{timestamp}.xlsx"
        backup_path = os.path.join(BACKUP_DIR, backup_name)

        shutil.copy2(ALERTAS_FILE, backup_path)
        print("Backup creado en:", backup_path)

    wb = Workbook()
    ws = wb.active
    ws.title = ALERTAS_SHEET

    ws.append([
        "ID",
        "Numero_Parte_o_Area",
        "Numero_Alerta",
        "Descripcion",
        "Fecha_Generacion",
        "Vigente_Hasta"
    ])

    for i, alerta in enumerate(data, start=1):
        ws.append([
            i,
            alerta["numero_parte_area"],
            alerta["numero_alerta"],
            alerta["descripcion"],
            alerta["fecha_generacion"],
            alerta["vigente_hasta"]
        ])

    wb.save(ALERTAS_FILE)

    print("Guardado exitoso.")
    return True
